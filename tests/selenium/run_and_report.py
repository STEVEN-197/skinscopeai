"""Run Selenium suite and produce an Excel results report next to the HTML report."""
import subprocess, sys, json, pathlib, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = pathlib.Path(__file__).resolve().parents[1]
REPORTS = ROOT / "reports"
EXCEL = ROOT / "excel-reports"
REPORTS.mkdir(parents=True, exist_ok=True)
EXCEL.mkdir(parents=True, exist_ok=True)

junit = REPORTS / "selenium_junit.xml"
cmd = [sys.executable, "-m", "pytest", "tests/selenium/scripts",
       f"--junitxml={junit}",
       f"--html={REPORTS/'selenium_report.html'}", "--self-contained-html"]
print(">>", " ".join(cmd))
result = subprocess.run(cmd, cwd=ROOT.parent)

# Parse junit
import xml.etree.ElementTree as ET
wb = Workbook(); ws = wb.active; ws.title = "Selenium Results"
hdr_fill = PatternFill("solid", start_color="1E40AF")
hdr_font = Font(bold=True, color="FFFFFF")
for col, h in enumerate(["Test", "Class", "Status", "Time (s)", "Message"], 1):
    c = ws.cell(row=1, column=col, value=h); c.fill = hdr_fill; c.font = hdr_font
total = passed = failed = skipped = 0
if junit.exists():
    tree = ET.parse(junit)
    for tc in tree.iter("testcase"):
        total += 1
        status = "PASS"; msg = ""
        if tc.find("failure") is not None: status = "FAIL"; msg = tc.find("failure").get("message", ""); failed += 1
        elif tc.find("error") is not None: status = "ERROR"; msg = tc.find("error").get("message", ""); failed += 1
        elif tc.find("skipped") is not None: status = "SKIP"; skipped += 1
        else: passed += 1
        ws.append([tc.get("name"), tc.get("classname"), status, float(tc.get("time", 0)), msg])

s2 = wb.create_sheet("Summary")
s2.append(["Generated", datetime.datetime.now().isoformat(timespec="seconds")])
s2.append(["Total", total]); s2.append(["Passed", passed])
s2.append(["Failed", failed]); s2.append(["Skipped", skipped])
s2.append(["Pass Rate", f"{(passed/total*100):.1f}%" if total else "n/a"])
for c in ws.column_dimensions: ws.column_dimensions[c].width = 28
out = EXCEL / "Selenium_Execution_Report.xlsx"
wb.save(out)
print(f"Excel report: {out}")
sys.exit(result.returncode)
