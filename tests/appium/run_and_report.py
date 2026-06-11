"""Run Appium suite and produce HTML + Excel reports."""
import subprocess, sys, pathlib, datetime, xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = pathlib.Path(__file__).resolve().parents[1]
REP = ROOT / "mobile-reports"; REP.mkdir(parents=True, exist_ok=True)
junit = REP / "appium_junit.xml"

cmd = [sys.executable, "-m", "pytest", "tests/appium/scripts",
       f"--junitxml={junit}",
       f"--html={REP/'appium_report.html'}", "--self-contained-html"]
print(">>", " ".join(cmd))
res = subprocess.run(cmd, cwd=ROOT.parent)

wb = Workbook(); ws = wb.active; ws.title = "Appium Results"
hf = PatternFill("solid", start_color="0EA5B7"); ft = Font(bold=True, color="FFFFFF")
for col, h in enumerate(["Test", "Class", "Status", "Time (s)", "Message"], 1):
    c = ws.cell(row=1, column=col, value=h); c.fill = hf; c.font = ft
total=passed=failed=skipped=0
if junit.exists():
    for tc in ET.parse(junit).iter("testcase"):
        total += 1
        status = "PASS"; msg = ""
        if tc.find("failure") is not None: status = "FAIL"; msg = tc.find("failure").get("message",""); failed += 1
        elif tc.find("error") is not None: status = "ERROR"; failed += 1
        elif tc.find("skipped") is not None: status = "SKIP"; skipped += 1
        else: passed += 1
        ws.append([tc.get("name"), tc.get("classname"), status, float(tc.get("time",0)), msg])
s2 = wb.create_sheet("Summary")
s2.append(["Generated", datetime.datetime.now().isoformat(timespec="seconds")])
for k,v in [("Total",total),("Passed",passed),("Failed",failed),("Skipped",skipped)]:
    s2.append([k,v])
out = ROOT / "mobile-reports" / "Appium_Execution_Report.xlsx"
wb.save(out); print(f"Excel: {out}")
sys.exit(res.returncode)
